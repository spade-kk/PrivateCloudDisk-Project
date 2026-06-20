"""
Excel 文件抽取器

支持: .xlsx, .xls
"""
from pathlib import Path

from openpyxl import load_workbook

from core.extractors.base import BaseExtractor, ExtractionResult, split_text, TextChunk
from core.config import settings


class XlsxExtractor(BaseExtractor):
    name = "xlsx_extractor"

    async def extract(self, file_id: str, path: Path) -> ExtractionResult:
        warnings: list[str] = []
        all_text_parts: list[str] = []
        chunks: list[TextChunk] = []

        try:
            workbook = load_workbook(path, read_only=True, data_only=True)

            for sheet in workbook.worksheets:
                rows_text: list[str] = []

                for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if row_index > 5000:
                        warnings.append(f"工作表 '{sheet.title}' 超过 5000 行，已截断")
                        break

                    values = [str(v) for v in row if v is not None and str(v).strip()]
                    if not values:
                        continue

                    row_text = f"{sheet.title} row {row_index}: " + " | ".join(values)
                    rows_text.append(row_text)
                    all_text_parts.append(row_text)

                    chunks.extend(
                        split_text(
                            text=row_text,
                            source_type="xlsx",
                            chunk_size=settings.chunk_size_chars,
                            overlap=settings.chunk_overlap_chars,
                            sheet=sheet.title,
                            row=row_index,
                        )
                    )

            workbook.close()
        except Exception as e:
            return ExtractionResult(
                file_id=file_id,
                extractor=self.name,
                warnings=[f"Excel 解析失败: {e}"],
            )

        text = "\n".join(all_text_parts)[:settings.content_max_chars]

        return ExtractionResult(
            file_id=file_id,
            extractor=self.name,
            extractor_version=self.version,
            content_text=text,
            chunks=chunks,
            tags=["table", "spreadsheet"],
            warnings=warnings,
        )